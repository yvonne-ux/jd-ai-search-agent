"""Workflow 5 — Longlist Compilation.

Aggregates candidate qualification summaries into a ranked, client-ready
longlist with market commentary, and exports it to Excel and Markdown.
Implements Prompt 4 from the developer brief.
"""
from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from agent.claude_client import MODEL_SONNET, ClaudeClient
from agent.logger import log_run
from agent.models import Longlist, QualificationSummary
from agent.prompts import fill_template, load_prompt

WORKFLOW = "longlist"
SYSTEM_FILE = "longlist_system.txt"
USER_FILE = "longlist_user.txt"

# The brief triggers this workflow once 15+ summaries are collected.
MIN_SUMMARIES = 15

_LONGLIST_DIR = Path(__file__).resolve().parent.parent / "data" / "longlists"


@dataclass
class LonglistContext:
    """The role context for a longlist."""

    role_title: str = ""
    client_type: str = ""
    location: str = ""
    must_have: str = ""

    @classmethod
    def from_intake_file(cls, path) -> "LonglistContext":
        """Build role context from an intake workflow JSON file."""
        data = json.loads(Path(path).read_text(encoding="utf-8"))
        brief = data.get("brief", data)
        return cls(
            role_title=str(brief.get("role_title", "") or "").strip(),
            client_type=str(
                brief.get("industry", "") or brief.get("client_name", "") or ""
            ).strip(),
            location=str(brief.get("location", "") or "").strip(),
            must_have=str(brief.get("must_have", "") or "").strip(),
        )


def load_qualification_summaries(directory) -> List[QualificationSummary]:
    """Load every qualification summary JSON file from a directory."""
    folder = Path(directory)
    if not folder.is_dir():
        return []

    summaries: List[QualificationSummary] = []
    for path in sorted(folder.glob("*.json")):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        record = data.get("qualification_summary", data)
        summaries.append(QualificationSummary.from_dict(record))
    return summaries


def compile_longlist(
    summaries: List[QualificationSummary],
    context: LonglistContext,
    client: ClaudeClient,
    *,
    max_tokens: int = 4096,
) -> Longlist:
    """Run Prompt 4 over the summaries and return a ranked Longlist."""
    system = load_prompt(SYSTEM_FILE)
    user = fill_template(
        load_prompt(USER_FILE),
        role_title=context.role_title or "(not stated)",
        client_type=context.client_type or "(not stated)",
        location=context.location or "(not stated)",
        must_have=context.must_have or "(not stated)",
        array_of_qualification_summaries=json.dumps(
            [s.to_dict() for s in summaries], indent=2, ensure_ascii=False
        ),
    )

    raw = client.complete_json(
        system=system,
        user=user,
        model=MODEL_SONNET,
        max_tokens=max_tokens,
    )
    log_run(
        WORKFLOW,
        MODEL_SONNET,
        {"context": asdict(context), "summary_count": len(summaries)},
        system,
        user,
        raw,
    )

    longlist = Longlist.from_dict(raw)
    # Re-sort by fit score and renumber, so rank is always consistent.
    longlist.entries.sort(key=lambda e: e.fit_score, reverse=True)
    for position, entry in enumerate(longlist.entries, 1):
        entry.rank = position
    return longlist


def format_longlist(longlist: Longlist) -> str:
    """Render the longlist as a readable block for consultant review."""
    if not longlist.entries:
        return "(no candidates on the longlist)"

    lines: List[str] = []
    for entry in longlist.entries:
        role = ", ".join(
            part for part in [entry.current_title, entry.current_company] if part
        )
        header = f"#{entry.rank}  {entry.candidate_name or 'Unknown'}"
        if role:
            header += f" — {role}"
        lines.append(header)
        lines.append(
            f"    Fit {entry.fit_score}/10 · Interest: "
            f"{entry.interest_level or 'n/a'}"
        )
        if entry.one_line_summary:
            lines.append(f"    {entry.one_line_summary}")
        if entry.recommended_next_step:
            lines.append(f"    Next step: {entry.recommended_next_step}")
        lines.append("")
    if longlist.search_commentary:
        lines.append("Search commentary:")
        lines.append(f"  {longlist.search_commentary}")
    return "\n".join(lines).rstrip()


def longlist_to_markdown(longlist: Longlist, context: LonglistContext) -> str:
    """Render the longlist as a client-ready Markdown document."""
    head = [
        f"# Longlist — {context.role_title or 'Role'}",
        "",
        f"**Client type:** {context.client_type or 'n/a'}  ",
        f"**Location:** {context.location or 'n/a'}  ",
        f"**Compiled:** {date.today().isoformat()}",
        "",
        "| Rank | Candidate | Current Role | Interest | Fit | Summary | Next Step |",
        "|-----:|-----------|--------------|----------|----:|---------|-----------|",
    ]
    for entry in longlist.entries:
        role = ", ".join(
            part for part in [entry.current_title, entry.current_company] if part
        )
        cells = [
            str(entry.rank),
            entry.candidate_name or "Unknown",
            role or "n/a",
            entry.interest_level or "n/a",
            f"{entry.fit_score}/10",
            entry.one_line_summary or "",
            entry.recommended_next_step or "",
        ]
        # Escape pipes so they do not break the Markdown table.
        head.append("| " + " | ".join(c.replace("|", "\\|") for c in cells) + " |")

    head += ["", "## Search Commentary", "",
             longlist.search_commentary or "(none provided)", ""]
    return "\n".join(head)


def longlist_to_excel(
    longlist: Longlist,
    context: LonglistContext,
    path: Path,
) -> Path:
    """Write the longlist to a colour-coded Excel workbook."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Longlist"

    headers = ["Rank", "Candidate", "Current Title", "Current Company",
               "Interest", "Fit Score", "One-Line Summary",
               "Recommended Next Step"]
    sheet.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1D9E75")
    for col, _ in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    for offset, entry in enumerate(longlist.entries, start=2):
        sheet.append([
            entry.rank,
            entry.candidate_name,
            entry.current_title,
            entry.current_company,
            entry.interest_level,
            entry.fit_score,
            entry.one_line_summary,
            entry.recommended_next_step,
        ])
        fit_cell = sheet.cell(row=offset, column=6)
        if entry.fit_score >= 8:
            fit_cell.fill = PatternFill(fill_type="solid", fgColor="DCFCE7")
        elif entry.fit_score >= 5:
            fit_cell.fill = PatternFill(fill_type="solid", fgColor="FEF9C3")
        else:
            fit_cell.fill = PatternFill(fill_type="solid", fgColor="FEE2E2")
        for col in (7, 8):
            sheet.cell(row=offset, column=col).alignment = Alignment(
                wrap_text=True, vertical="top")

    for col, width in zip("ABCDEFGH", [6, 22, 24, 22, 14, 10, 44, 32]):
        sheet.column_dimensions[col].width = width

    commentary = workbook.create_sheet("Commentary")
    commentary["A1"] = "Search Commentary"
    commentary["A1"].font = Font(bold=True)
    commentary["A2"] = longlist.search_commentary or "(none provided)"
    commentary["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    commentary.column_dimensions["A"].width = 100

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _slug(text: str) -> str:
    slug = re.sub(r"[^\w\s-]", "", text or "").strip().lower()
    slug = re.sub(r"[\s_-]+", "-", slug)
    return slug or "longlist"


def save_longlist(
    longlist: Longlist,
    context: LonglistContext,
    name: str,
    out_dir: Optional[Path] = None,
) -> Dict[str, Path]:
    """Write the longlist as JSON, Markdown, and Excel. Return the file paths."""
    target = Path(out_dir) if out_dir else _LONGLIST_DIR
    target.mkdir(parents=True, exist_ok=True)

    base = _slug(name)
    suffix = ""
    counter = 2
    while (target / f"{base}{suffix}.json").exists():
        suffix = f"-{counter}"
        counter += 1
    base += suffix

    json_path = target / f"{base}.json"
    json_path.write_text(
        json.dumps(
            {"role": asdict(context), "longlist": longlist.to_dict()},
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    md_path = target / f"{base}.md"
    md_path.write_text(longlist_to_markdown(longlist, context), encoding="utf-8")

    xlsx_path = longlist_to_excel(longlist, context, target / f"{base}.xlsx")

    return {"json": json_path, "markdown": md_path, "excel": xlsx_path}
