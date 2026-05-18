"""Unit tests for Workflow 5 — Longlist Compilation.

Covers the non-API parts: prompt loading, summary aggregation, role context,
longlist formatting, and the Markdown / Excel exporters. Run from the project
root:
    python -m unittest discover tests
"""
import json
import tempfile
import unittest
from pathlib import Path

import openpyxl

from agent.models import Longlist, LonglistEntry, QualificationSummary
from agent.prompts import load_prompt
from workflows.longlist import (
    SYSTEM_FILE,
    USER_FILE,
    LonglistContext,
    format_longlist,
    load_qualification_summaries,
    longlist_to_excel,
    longlist_to_markdown,
    save_longlist,
)


def _sample_longlist() -> Longlist:
    return Longlist(
        entries=[
            LonglistEntry(
                rank=1,
                candidate_name="David Wong",
                current_title="CFO",
                current_company="SeaFreight Holdings",
                interest_level="Medium",
                fit_score=9,
                one_line_summary="Regional CFO with IPO appetite.",
                recommended_next_step="Consultant call this week.",
            ),
            LonglistEntry(
                rank=2,
                candidate_name="Lena Ortega",
                current_title="Finance Director",
                current_company="PortLink",
                interest_level="Low",
                fit_score=4,
                one_line_summary="Strong FD but no IPO exposure.",
                recommended_next_step="Hold for now.",
            ),
        ],
        search_commentary="Healthy response rate from logistics CFOs.",
    )


class PromptTests(unittest.TestCase):
    def test_longlist_prompts_load(self):
        system = load_prompt(SYSTEM_FILE)
        self.assertIn("JonDavidson", system)
        self.assertIn("search_commentary", system)

    def test_user_template_has_expected_placeholders(self):
        template = load_prompt(USER_FILE)
        for placeholder in ("{{role_title}}",
                            "{{array_of_qualification_summaries}}"):
            self.assertIn(placeholder, template)


class LoadSummariesTests(unittest.TestCase):
    def test_loads_summaries_from_directory(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            (folder / "a.json").write_text(json.dumps({
                "qualification_summary": {"candidate_name": "David Wong"},
            }), encoding="utf-8")
            (folder / "b.json").write_text(json.dumps({
                "qualification_summary": {"candidate_name": "Lena Ortega"},
            }), encoding="utf-8")
            summaries = load_qualification_summaries(folder)
            self.assertEqual(len(summaries), 2)
            self.assertEqual(summaries[0].candidate_name, "David Wong")

    def test_missing_directory_returns_empty(self):
        self.assertEqual(load_qualification_summaries("/no/such/dir"), [])

    def test_skips_unparseable_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp)
            (folder / "good.json").write_text(json.dumps({
                "qualification_summary": {"candidate_name": "David Wong"},
            }), encoding="utf-8")
            (folder / "bad.json").write_text("{not json", encoding="utf-8")
            summaries = load_qualification_summaries(folder)
            self.assertEqual(len(summaries), 1)


class LonglistContextTests(unittest.TestCase):
    def test_from_intake_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "criteria.json"
            path.write_text(json.dumps({
                "brief": {
                    "role_title": "CFO",
                    "industry": "Logistics",
                    "location": "Singapore",
                    "must_have": "IPO track record",
                },
            }), encoding="utf-8")
            context = LonglistContext.from_intake_file(path)
            self.assertEqual(context.role_title, "CFO")
            self.assertEqual(context.client_type, "Logistics")


class FormatLonglistTests(unittest.TestCase):
    def test_format_includes_entries_and_commentary(self):
        text = format_longlist(_sample_longlist())
        self.assertIn("#1", text)
        self.assertIn("David Wong", text)
        self.assertIn("Fit 9/10", text)
        self.assertIn("Search commentary", text)

    def test_format_handles_empty_longlist(self):
        self.assertEqual(
            format_longlist(Longlist()), "(no candidates on the longlist)"
        )


class MarkdownExportTests(unittest.TestCase):
    def test_markdown_has_table_and_commentary(self):
        context = LonglistContext(role_title="CFO", client_type="Logistics")
        md = longlist_to_markdown(_sample_longlist(), context)
        self.assertIn("# Longlist — CFO", md)
        self.assertIn("| Rank |", md)
        self.assertIn("David Wong", md)
        self.assertIn("## Search Commentary", md)

    def test_markdown_escapes_pipes(self):
        longlist = Longlist(entries=[
            LonglistEntry(rank=1, candidate_name="A|B", fit_score=5),
        ])
        md = longlist_to_markdown(longlist, LonglistContext())
        self.assertIn("A\\|B", md)


class ExcelExportTests(unittest.TestCase):
    def test_excel_file_is_written_and_readable(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "longlist.xlsx"
            context = LonglistContext(role_title="CFO")
            longlist_to_excel(_sample_longlist(), context, path)
            self.assertTrue(path.exists())
            workbook = openpyxl.load_workbook(path)
            self.assertIn("Longlist", workbook.sheetnames)
            self.assertIn("Commentary", workbook.sheetnames)
            sheet = workbook["Longlist"]
            self.assertEqual(sheet.cell(row=1, column=1).value, "Rank")
            self.assertEqual(sheet.cell(row=2, column=2).value, "David Wong")


class SaveLonglistTests(unittest.TestCase):
    def test_save_writes_three_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            context = LonglistContext(role_title="CFO")
            paths = save_longlist(
                _sample_longlist(), context, "acme-cfo", out_dir=Path(tmp)
            )
            for key in ("json", "markdown", "excel"):
                self.assertTrue(paths[key].exists(), msg=key)

    def test_save_does_not_clobber(self):
        with tempfile.TemporaryDirectory() as tmp:
            context = LonglistContext(role_title="CFO")
            first = save_longlist(Longlist(), context, "acme-cfo",
                                  out_dir=Path(tmp))
            second = save_longlist(Longlist(), context, "acme-cfo",
                                   out_dir=Path(tmp))
            self.assertNotEqual(first["json"], second["json"])
            self.assertEqual(second["json"].name, "acme-cfo-2.json")


if __name__ == "__main__":
    unittest.main()
